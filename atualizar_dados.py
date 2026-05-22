# -*- coding: utf-8 -*-
import openpyxl
import json
import os
import sys

# Forçar saída do terminal em UTF-8 no Windows
if sys.platform.startswith('win'):
    import codecs
    sys.stdout = codecs.getwriter('utf-8')(sys.stdout.detach())

# Definir caminhos relativos ao local do script
base_dir = os.path.dirname(os.path.abspath(__file__))
excel_name = "Tabela Revisão Programada FIAT Março 2026.xlsx"
excel_path = os.path.join(base_dir, excel_name)
json_path = os.path.join(base_dir, "fiat_revisoes_data.json")
js_path = os.path.join(base_dir, "fiat_data.js")

# Mapeamento de nomes para modelos Titano e Scudo conforme solicitação do usuário
mapeamento_nomes = {
    "TITANO 2.2D AT": "TITANO 2.2D AT (Nova/Argentina/8AP)",
    "TITANO 2.2D MT": "TITANO 2.2D MT (Nova/Argentina/8AP)",
    "TITANO AUTOMÁTICO": "TITANO 2.2D AT (Antiga/Uruguaia/9VC)",
    "TITANO MANUAL": "TITANO 2.2D MT (Antiga/Uruguaia/9VC)",
    "SCUDO": "SCUDO 1.5 (Antiga)",
    "SCUDO 2.2D": "SCUDO 2.2 (Nova)"
}





print("==============================================================")
print("     EXTRAÇÃO DE DADOS DE REVISÃO FIAT (MÚLTIPLOS MODELOS)")
print("==============================================================")

if not os.path.exists(excel_path):
    print(f"\n[ERRO] O arquivo '{excel_name}' não foi encontrado na pasta!")
    print(f"Caminho esperado: {excel_path}")
    print("Por favor, coloque a planilha Excel correta nesta mesma pasta.")
    sys.exit(1)

print(f"\nLendo arquivo Excel: {excel_name}...")
try:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
except Exception as e:
    print(f"\n[ERRO] Não foi possível abrir o arquivo Excel: {e}")
    sys.exit(1)

sheet_names = wb.sheetnames
data = {
    "modelos": {}
}

# 1. Carregar preços nacionais da primeira aba (PREÇO NACIONAL)
precos_nacionais = {}
sheet_precos = wb[sheet_names[0]]
print(f"Processando aba resumo: {sheet_names[0]}")

# Encontrar linha de cabeçalho na aba PREÇO NACIONAL
header_row_idx = None
for r in range(1, 10):
    val = sheet_precos.cell(row=r, column=2).value
    if val and any(x in str(val).lower() for x in ['modelo', 'ano']):
        header_row_idx = r
        break

if header_row_idx:
    col_headers = {}
    for c in range(3, 20):
        val = sheet_precos.cell(row=header_row_idx, column=c).value
        if val:
            col_headers[c] = str(val).strip()
    
    for r in range(header_row_idx + 1, sheet_precos.max_row + 1):
        modelo_nome = sheet_precos.cell(row=r, column=2).value
        if not modelo_nome or str(modelo_nome).strip() == "" or "Preço" in str(modelo_nome) or "a cada" in str(modelo_nome):
            continue
        modelo_nome = str(modelo_nome).strip()
        if modelo_nome in mapeamento_nomes:
            modelo_nome = mapeamento_nomes[modelo_nome]
        
        revisoes = {}
        for c, h in col_headers.items():
            val = sheet_precos.cell(row=r, column=c).value
            if val is not None:
                try:
                    revisoes[h] = float(val)
                except ValueError:
                    revisoes[h] = val
        
        total_acumulado = sheet_precos.cell(row=r, column=13).value
        if total_acumulado is not None:
            try:
                total_acumulado = float(total_acumulado)
            except:
                pass
                
        precos_nacionais[modelo_nome] = {
            "revisoes": revisoes,
            "total_acumulado": total_acumulado
        }
    print(f"  -> Preços nacionais extraídos para {len(precos_nacionais)} modelos.")
else:
    print("  [AVISO] Cabeçalho de Preço Nacional não encontrado.")

data["precos_nacionais"] = precos_nacionais

# 2. Processar cada aba de veículo individual (incluindo abas com múltiplos blocos)
for sheet_name in sheet_names[1:]:
    sheet = wb[sheet_name]
    
    # Encontrar todas as linhas de cabeçalho do bloco (ITENS DE SUBSTITUIÇÃO OBRIGATÓRIA)
    header_rows = []
    for r in range(1, sheet.max_row + 1):
        val = sheet.cell(row=r, column=2).value
        if val and any(x in str(val).lower() for x in ['itens de substitui', 'itens de substituic', 'itens de substituição']):
            header_rows.append(r)
            
    print(f"\nProcessando aba: '{sheet_name}' (contém {len(header_rows)} modelo(s))...")
    
    for idx_block, peças_header_row in enumerate(header_rows):
        # Determinar nome do modelo para este bloco
        nome_modelo = None
        for r_subir in range(peças_header_row - 1, 0, -1):
            val1 = sheet.cell(row=r_subir, column=1).value
            if val1 is not None and str(val1).strip() != "":
                nome_modelo = str(val1).strip()
                if nome_modelo in mapeamento_nomes:
                    nome_modelo = mapeamento_nomes[nome_modelo]
                break
        if not nome_modelo:
            nome_modelo = sheet_name
            
        print(f"  -> Bloco {idx_block + 1}: '{nome_modelo}'")
        
        # Encontrar custo_start_col
        custo_start_col = None
        max_cols = sheet.max_column
        for c in range(5, max_cols + 1):
            val = sheet.cell(row=peças_header_row, column=c).value
            if val is not None:
                val_str = str(val).strip()
                if val_str.isdigit() and int(val_str) >= 1000:
                    custo_start_col = c
                    break
        if not custo_start_col:
            # Fallback para detectar coluna de custos por textos (ex: 'km')
            for c in range(5, max_cols + 1):
                val = sheet.cell(row=peças_header_row, column=c).value
                if val and "km" in str(val).lower():
                    custo_start_col = c
                    break
                    
        if not custo_start_col:
            print(f"    [ERRO] Não foi possível achar o início dos custos para o bloco '{nome_modelo}'. Pulando...")
            continue
            
        num_revisoes = custo_start_col - 5
        
        revisoes_lista = []
        kms_lista = []
        custos_revisoes_totais = []
        
        for i in range(num_revisoes):
            qty_col = 5 + i
            cost_col = custo_start_col + i
            
            rev_val = sheet.cell(row=peças_header_row, column=qty_col).value
            rev_name = str(rev_val).strip() if rev_val is not None else f"{i+1}ª"
            revisoes_lista.append(rev_name)
            
            km_val = sheet.cell(row=peças_header_row, column=cost_col).value
            km_name = f"{int(km_val)} km" if isinstance(km_val, (int, float)) else str(km_val).strip()
            kms_lista.append(km_name)
            
            tot_val = sheet.cell(row=peças_header_row - 1, column=cost_col).value
            if tot_val is None:
                tot_val = sheet.cell(row=peças_header_row - 2, column=cost_col).value
                
            custo_total = float(tot_val) if isinstance(tot_val, (int, float)) else 0.0
            custos_revisoes_totais.append(custo_total)
            
        # Determinar linha de fim do bloco atual (antes do próximo cabeçalho ou até uma linha vazia/total)
        end_row = sheet.max_row
        next_headers = [r for r in header_rows if r > peças_header_row]
        if next_headers:
            end_row = next_headers[0] - 1
            
        itens = []
        for r in range(peças_header_row + 1, end_row + 1):
            num_item = sheet.cell(row=r, column=1).value
            if num_item is None or str(num_item).strip() == "" or "total" in str(num_item).lower():
                desc = sheet.cell(row=r, column=2).value
                if not desc or str(desc).strip() == "" or "total" in str(desc).lower():
                    break
            
            desc = str(sheet.cell(row=r, column=2).value).strip()
            pn = str(sheet.cell(row=r, column=3).value).strip() if sheet.cell(row=r, column=3).value is not None else ""
            preco_unit = sheet.cell(row=r, column=4).value
            
            if preco_unit is not None:
                try:
                    preco_unit = float(preco_unit)
                except ValueError:
                    pass
                    
            trocas = {}
            custos_itens = {}
            for i in range(num_revisoes):
                qty_col = 5 + i
                cost_col = custo_start_col + i
                rev_name = revisoes_lista[i]
                
                qty_val = sheet.cell(row=r, column=qty_col).value
                cost_val = sheet.cell(row=r, column=cost_col).value
                
                if qty_val is not None and str(qty_val).strip() != "" and str(qty_val).strip().lower() != 'nan':
                    try:
                        trocas[rev_name] = float(qty_val) if '.' in str(qty_val) else int(qty_val)
                    except:
                        trocas[rev_name] = str(qty_val).strip()
                        
                if cost_val is not None and str(cost_val).strip() != "" and str(cost_val).strip().lower() != 'nan':
                    try:
                        custos_itens[rev_name] = float(cost_val)
                    except:
                        custos_itens[rev_name] = str(cost_val).strip()
                        
            tipo = "peça"
            if "mão-de-obra" in desc.lower() or "mão de obra" in desc.lower() or "mo fiat" in pn.lower() or "tempo padrão" in desc.lower() or "total de mão de obra" in desc.lower():
                tipo = "serviço"
                if preco_unit == 342.0:
                    preco_unit = 349.0
                
                # Recalcular custos com base no novo valor de mão de obra
                for r_name in trocas:
                    try:
                        custos_itens[r_name] = round(float(trocas[r_name]) * preco_unit, 2)
                    except:
                        pass
                
            itens.append({
                "nome": desc,
                "pn": pn,
                "preco_unitario": preco_unit,
                "tipo": tipo,
                "trocas": trocas,
                "custos": custos_itens
            })
            
        # Aplicar regras customizadas de óleo para Ducato e Ducato X250 2.2D
        if nome_modelo == "DUCATO":
            for item in itens:
                name_lower = item["nome"].lower()
                if item["tipo"] == "peça" and any(x in name_lower for x in ["óleo", "oleo"]) and "motor" in name_lower and not any(x in name_lower for x in ["filtro", "filtrante"]):
                    item["nome"] = "5W30"
                    item["pn"] = "K68231015LA"
                    item["preco_unitario"] = 91.54
                    for r_name in list(item["trocas"].keys()):
                        try:
                            val = float(item["trocas"][r_name])
                            if val > 0:
                                item["trocas"][r_name] = 5.6
                        except:
                            pass
                    for r_name in list(item["custos"].keys()):
                        try:
                            val = float(item["custos"][r_name])
                            if val > 0:
                                item["custos"][r_name] = round(5.6 * 91.54, 2)
                        except:
                            pass
        elif nome_modelo == "DUCATO X250 2.2D":
            for item in itens:
                name_lower = item["nome"].lower()
                if item["tipo"] == "peça" and any(x in name_lower for x in ["óleo", "oleo"]) and "motor" in name_lower and not any(x in name_lower for x in ["filtro", "filtrante"]):
                    item["nome"] = "5W30"
                    item["pn"] = "K68231015LA"
                    item["preco_unitario"] = 91.54
                    for r_name in list(item["trocas"].keys()):
                        try:
                            val = float(item["trocas"][r_name])
                            if val > 0:
                                item["trocas"][r_name] = 6.0
                        except:
                            pass
                    for r_name in list(item["custos"].keys()):
                        try:
                            val = float(item["custos"][r_name])
                            if val > 0:
                                item["custos"][r_name] = round(6.0 * 91.54, 2)
                        except:
                            pass
            
        # Recalcular custos totais de cada revisão baseados nos novos custos recalculados
        custos_revisoes_totais = []
        for i in range(num_revisoes):
            rev_name = revisoes_lista[i]
            soma_rev = 0.0
            for item in itens:
                if rev_name in item["custos"] and isinstance(item["custos"][rev_name], (int, float)):
                    soma_rev += item["custos"][rev_name]
            custos_revisoes_totais.append(round(soma_rev, 2))
            
        data["modelos"][nome_modelo] = {
            "modelo": nome_modelo,
            "revisoes": revisoes_lista,
            "quilometragens": kms_lista,
            "custos_totais": custos_revisoes_totais,
            "itens": itens
        }
        
        # Cruzar e atualizar precos_nacionais para este modelo específico
        matched_key = None
        for pk in precos_nacionais.keys():
            if pk.lower().strip() == nome_modelo.lower().strip():
                matched_key = pk
                break
                
        if matched_key:
            new_revisoes = {}
            for idx, r_name in enumerate(revisoes_lista):
                pk_rev_keys = list(precos_nacionais[matched_key]["revisoes"].keys())
                if idx < len(pk_rev_keys):
                    rev_key = pk_rev_keys[idx]
                    new_revisoes[rev_key] = custos_revisoes_totais[idx]
            
            precos_nacionais[matched_key]["revisoes"] = new_revisoes
            # Recalcular total acumulado (soma das 10 primeiras revisões)
            precos_nacionais[matched_key]["total_acumulado"] = round(sum(custos_revisoes_totais[:10]), 2)
            print(f"    [Sucesso] Mapeado com resumo nacional para '{nome_modelo}'")
        else:
            print(f"    [Aviso] '{nome_modelo}' não foi associado a um modelo na aba resumo nacional.")

# Salvar o JSON consolidado
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)
print(f"\nJSON de dados gerado em: {json_path}")

# Salvar o JS global exposto
with open(js_path, 'w', encoding='utf-8') as f:
    f.write("// Dados consolidados de revisões Fiat\n")
    f.write("window.fiatData = ")
    json.dump(data, f, ensure_ascii=False, indent=2)
    f.write(";\n")
print(f"JS de dados gerado com sucesso em: {js_path}")
print("\nProcesso concluído com sucesso!")
print("==============================================================")
